"""Shatters the Ames Housing dataset into 3 fake "source systems" mirroring what a bank's
data estate actually looks like — meeting data where it lives, per the dlt ingestion demo (see
README's Data Ingestion section). Every generated field still comes from Ames' own columns or
a deterministic seeded RNG; no Faker-fabricated names/addresses, so this doesn't quietly
undercut the "Zero-scrape ingestion" PRD line the rest of the app already satisfies.

Sources (each owns a different subset of the book, mirroring which system would plausibly hold
which data in a real bank):
- core_banking: every property (every mortgaged asset has a loan account) — CSV, SFTP-shaped:
  renamed columns, zero-padded string PIDs, DD/MM/YYYY dates, a few deliberately malformed rows.
- valuation_vendor: a ~40% subset (only some properties were externally revalued) — served as
  paginated JSON by vendor_api/main.py.
- valuations_team: a ~25% subset (only some had a manual valuations-team review) — xlsx with a
  merged two-row header and a free-text notes column, a few malformed rows.

Run with --drift to additionally drop a new core_banking batch that introduces one new column
(disbursement_channel) on a small subset of already-loaded PIDs — the schema-drift demo beat:
dlt evolves the destination schema automatically, nothing breaks.
"""
import argparse
import json
import sys
from datetime import date, timedelta
from pathlib import Path

import numpy as np
import openpyxl

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "backend"))

from train_model import load_frame  # noqa: E402

CORE_BANKING_DIR = ROOT / "ingestion" / "dropzone" / "core_banking"
VALUATIONS_TEAM_DIR = ROOT / "ingestion" / "dropzone" / "valuations_team"
VENDOR_SEED_PATH = ROOT / "vendor_api" / "records_seed.json"

SEED = 20260710  # distinct from seed_db.py's 20260709 — a different simulated data domain
AS_OF = date(2026, 7, 1)


def zero_pad(pid: int) -> str:
    return str(pid).zfill(12)


def generate_core_banking(df, rng) -> list[dict]:
    ratios = rng.uniform(0.55, 0.88, size=len(df))
    rows = []
    for i, (_, r) in enumerate(df.iterrows()):
        pid = int(r["pid"])
        loan_bal = round(float(ratios[i]) * float(r["saleprice"]), 2)
        orig_dt = AS_OF - timedelta(days=int(rng.integers(30, 3650)))
        rows.append({
            "acct_pid": zero_pad(pid),
            "loan_bal": loan_bal,
            "orig_dt": orig_dt.strftime("%d/%m/%Y"),
            "last_updated": orig_dt.isoformat(),
        })

    # Deliberately malformed rows for the quarantine demo — mutate a handful in place.
    bad_idx = rng.choice(len(rows), size=min(8, len(rows)), replace=False)
    for j, idx in enumerate(bad_idx):
        if j % 2 == 0:
            rows[idx]["loan_bal"] = -abs(rows[idx]["loan_bal"])  # negative_loan_balance
        else:
            rows[idx]["acct_pid"] = "ACCT-" + rows[idx]["acct_pid"][-6:]  # pid_unparseable
    return rows


def write_core_banking_csv(rows: list[dict], filename: str) -> None:
    import csv
    CORE_BANKING_DIR.mkdir(parents=True, exist_ok=True)
    path = CORE_BANKING_DIR / filename
    fieldnames = ["acct_pid", "loan_bal", "orig_dt", "last_updated"]
    if any("disbursement_channel" in r for r in rows):
        fieldnames.append("disbursement_channel")
    with open(path, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)
    print(f"  wrote {path} ({len(rows)} rows)")


def generate_valuation_vendor(df, rng) -> list[dict]:
    subset = df.sample(frac=0.40, random_state=int(rng.integers(0, 2**31))).copy()
    records = []
    for _, r in subset.iterrows():
        noise = rng.uniform(0.92, 1.08)
        records.append({
            "pid": int(r["pid"]),
            "neighborhood": r["neighborhood"],
            "bldg_type": r["bldg_type"],
            "house_style": r["house_style"],
            "year_built": int(r["year_built"]),
            "overall_qual": int(r["overall_qual"]),
            "overall_cond": int(r["overall_cond"]),
            "gr_liv_area": int(r["gr_liv_area"]),
            "appraised_value": round(float(r["saleprice"]) * noise, 2),
            "last_updated": (AS_OF - timedelta(days=int(rng.integers(1, 400)))).isoformat(),
        })
    # A few malformed records — missing/non-numeric fields, never silently dropped downstream.
    bad_idx = rng.choice(len(records), size=min(5, len(records)), replace=False)
    for j, idx in enumerate(bad_idx):
        if j % 2 == 0:
            records[idx]["pid"] = None
        else:
            records[idx]["appraised_value"] = "TBD"
    return records


def generate_valuations_team(df, rng) -> list[dict]:
    subset = df.sample(frac=0.25, random_state=int(rng.integers(0, 2**31))).copy()
    reviewers = ["A. Reid", "K. Brown", "T. Facey"]
    notes_pool = [
        "Roof re-shingled 2019, no visible deferred maintenance.",
        "Kitchen renovation not yet reflected in comparable set.",
        "Boundary dispute flagged by neighbor — see file note.",
        "Standard drive-by review, no material findings.",
        "",
    ]
    rows = []
    for _, r in subset.iterrows():
        rows.append({
            "pid": int(r["pid"]),
            "valuation_figure": round(float(r["saleprice"]) * rng.uniform(0.95, 1.05), 2),
            "notes": notes_pool[int(rng.integers(0, len(notes_pool)))],
            "reviewed_by": reviewers[int(rng.integers(0, len(reviewers)))],
            "last_updated": (AS_OF - timedelta(days=int(rng.integers(1, 400)))).isoformat(),
        })
    bad_idx = rng.choice(len(rows), size=min(5, len(rows)), replace=False)
    for j, idx in enumerate(bad_idx):
        if j % 2 == 0:
            rows[idx]["pid"] = None
        else:
            rows[idx]["valuation_figure"] = "pending"
    return rows


def write_valuations_team_xlsx(rows: list[dict], filename: str) -> None:
    """Two-row header (a merged title, then real column names) — real spreadsheets handed
    off between teams are rarely a clean single-row header."""
    VALUATIONS_TEAM_DIR.mkdir(parents=True, exist_ok=True)
    columns = ["pid", "valuation_figure", "notes", "reviewed_by", "last_updated"]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Valuations"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    ws.cell(row=1, column=1, value="Valuations Team — Manual Review Log")
    for col_idx, name in enumerate(columns, start=1):
        ws.cell(row=2, column=col_idx, value=name)
    for row_idx, row in enumerate(rows, start=3):
        for col_idx, name in enumerate(columns, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(name))
    path = VALUATIONS_TEAM_DIR / filename
    wb.save(path)
    print(f"  wrote {path} ({len(rows)} rows)")


def write_vendor_seed(records: list[dict]) -> None:
    VENDOR_SEED_PATH.parent.mkdir(parents=True, exist_ok=True)
    VENDOR_SEED_PATH.write_text(json.dumps(records, indent=2, default=str))
    print(f"  wrote {VENDOR_SEED_PATH} ({len(records)} records)")


def generate_drift_batch(df, rng) -> list[dict]:
    """A small follow-up core_banking batch introducing a new column
    (`disbursement_channel`) on a handful of already-loaded PIDs — the schema-drift beat."""
    subset = df.sample(n=min(50, len(df)), random_state=int(rng.integers(0, 2**31)))
    channels = ["branch", "online", "correspondent"]
    rows = []
    for _, r in subset.iterrows():
        pid = int(r["pid"])
        orig_dt = AS_OF + timedelta(days=1)
        rows.append({
            "acct_pid": zero_pad(pid),
            "loan_bal": round(float(r["saleprice"]) * rng.uniform(0.55, 0.88), 2),
            "orig_dt": orig_dt.strftime("%d/%m/%Y"),
            "last_updated": orig_dt.isoformat(),
            "disbursement_channel": channels[int(rng.integers(0, len(channels)))],
        })
    return rows


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--drift", action="store_true",
                        help="also drop a follow-up core_banking batch with a new column")
    args = parser.parse_args()

    df = load_frame()
    rng = np.random.default_rng(SEED)

    print("Generating core_banking CSV drop...")
    write_core_banking_csv(generate_core_banking(df, rng), "loans_20260701.csv")

    print("Generating valuation_vendor API seed...")
    write_vendor_seed(generate_valuation_vendor(df, rng))

    print("Generating valuations_team spreadsheet...")
    write_valuations_team_xlsx(generate_valuations_team(df, rng), "valuations_report.xlsx")

    if args.drift:
        print("Generating schema-drift core_banking batch (adds disbursement_channel)...")
        write_core_banking_csv(generate_drift_batch(df, rng), "loans_20260702_drift.csv")

    print("Done. Re-run `POST /api/v1/ingestion/sync` (or the Data Ingestion tab) to pick these up.")


if __name__ == "__main__":
    main()
