"""dlt source: "the valuations team" spreadsheet — an xlsx with merged header cells, a
free-text notes column, and a few deliberately malformed rows, dropped by
scripts/generate_ingestion_sources.py into ingestion/dropzone/valuations_team/.
"""
from pathlib import Path

import dlt
import openpyxl

from ..mapping import validate_valuations_team

DROPZONE = Path(__file__).resolve().parents[1] / "dropzone" / "valuations_team"


def _read_rows() -> list[dict]:
    """The workbook's header spans two rows (a merged title row, then real column names) —
    real spreadsheets handed off between teams are rarely a clean single-row header."""
    if not DROPZONE.exists():  # fresh checkout, generator hasn't run yet — sync 0 rows, don't crash
        return []
    rows = []
    for path in sorted(DROPZONE.glob("*.xlsx")):
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        header = [str(c.value).strip() if c.value is not None else "" for c in ws[2]]
        for row_cells in ws.iter_rows(min_row=3):
            values = [c.value for c in row_cells]
            if all(v is None for v in values):
                continue
            rows.append(dict(zip(header, values)))
    return rows


@dlt.source(name="valuations_team")
def valuations_team_source():
    clean_rows, rejected_rows = [], []
    for row in _read_rows():
        mapped, reason = validate_valuations_team(row)
        if mapped is not None:
            clean_rows.append(mapped)
        else:
            reason_code, reason_detail = reason
            rejected_rows.append({
                "raw_record": {k: (v if v is None or isinstance(v, (str, int, float)) else str(v))
                              for k, v in row.items()},
                "reason_code": reason_code, "reason_detail": reason_detail,
            })

    @dlt.resource(name="valuations_team_clean", write_disposition="merge", primary_key="pid")
    def clean(last_updated=dlt.sources.incremental("last_updated", initial_value="")):
        yield clean_rows

    @dlt.resource(name="valuations_team_rejected", write_disposition="append",
                 columns={"raw_record": {"data_type": "json"}})
    def rejected():
        yield rejected_rows

    return clean, rejected
